# Datenschmiede test data generator — Python runtime.
#
# Invoked by the extension host with the path of a plan JSON file (see
# src/project/run.ts, where the plan is built from the .tdproject/.td/.lkp/
# .tdgen files). Produces synthetic records for every table of the plan —
# heavily vectorized via numpy/pandas so that even large data volumes are
# generated quickly — and writes them as CSV, Excel, JSON or XML according to
# the table's output configuration.
#
# Flow:
#   1. Sort the tables topologically (foreign key and generator references),
#      then the columns of each table — column by column, as far as the
#      dependencies (e.g. combine templates) allow.
#   2. Determine the row count: primary tables from their fixed count,
#      referenced tables from the cardinality per record of the referenced
#      table (the driving FK column is produced along the way, via
#      numpy.repeat).
#   3. Fill every column with its generator (built-in ones directly here,
#      custom ones from the Python code shipped in the plan).
#   4. Write the file in the configured format (see write_output) and resolve
#      the file name from its {…} template.
#
# Progress and result are reported as JSON lines on stdout (events: start /
# table_start / table_done / done / error) — the extension host translates them
# into the VS Code progress indicator.

import json
import re
import sys
import traceback
from datetime import datetime, timedelta
from xml.sax.saxutils import escape as xml_escape, quoteattr as xml_quoteattr

# The event protocol and the log output are UTF-8, independently of the
# platform: with a piped stdout Python otherwise falls back to the locale
# encoding (cp1252 on a German Windows), which turns every umlaut in a preview,
# a log line or a traceback into garbage on the extension host — which decodes
# UTF-8. Only the standard streams are affected here; the generated file keeps
# the encoding configured for the table (see write_csv/write_json/write_xml).
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass


def emit(event, **payload):
    """Writes one event of the stdout protocol read by the extension host."""
    print(json.dumps({"event": event, **payload}, ensure_ascii=False), flush=True)


def fail(message, **payload):
    """Reports an error event and terminates the run."""
    emit("error", message=str(message), **payload)
    sys.exit(2)


# Check dependencies early and all at once, so the message in the extension
# host can offer a concrete installation hint.
_missing = []
try:
    import numpy as np
except ImportError:
    _missing.append("numpy")
try:
    import pandas as pd
except ImportError:
    _missing.append("pandas")
if _missing:
    emit("error", code="missing-packages", packages=_missing,
         message="Missing Python packages: " + ", ".join(_missing))
    sys.exit(3)

import csv as csv_module

RNG = np.random.default_rng()
RUN_DT = datetime.now()

_faker_instances = {}


def get_faker(locale):
    """Faker instance per locale (imported lazily, with a clear error message)."""
    key = locale or "en_US"
    if key not in _faker_instances:
        try:
            from faker import Faker
        except ImportError:
            emit("error", code="missing-packages", packages=["faker"],
                 message="Missing Python package: faker")
            sys.exit(3)
        _faker_instances[key] = Faker(key)
    return _faker_instances[key]


class Context:
    """The `ctx` object handed to custom generators (see the .tdgen editor)."""

    def __init__(self, runner, table):
        self.rng = RNG
        self.np = np
        self.pd = pd
        self._runner = runner
        self._table = table

    def faker(self, locale="en_US"):
        return get_faker(locale)

    def log(self, *args):
        """
        Writes a message to the run log (the "Datenschmiede" output channel in
        VS Code). Deliberately provided instead of print(): stdout is reserved
        for the JSON event protocol — raw print() lines are discarded there
        (print(..., file=sys.stderr) also ends up in the output channel).
        """
        emit("log", table=self._table["label"], message=" ".join(str(a) for a in args))

    def column(self, name):
        """Already generated values of another column of this table."""
        data = self._runner.data.get(self._table["label"], {})
        if name not in data:
            raise RuntimeError(
                f'ctx.column("{name}"): column is not generated yet (or does not exist) '
                f'in table {self._table["label"]}'
            )
        return data[name]

    def table(self, label, column):
        """Already generated values of a column of another table in the plan."""
        data = self._runner.data.get(label)
        if data is None or column not in data:
            raise RuntimeError(f'ctx.table("{label}", "{column}"): values are not available (yet)')
        return data[column]

    def related(self, fk_path, column):
        """
        Row-accurate values of a referenced (parent) table: for every record of
        this table the value of `column` from EXACTLY the record its FK column
        points at — a join over fk_table/fk_column rather than a random sample
        (ctx.table).

        `fk_path` may also be a dot-separated path across SEVERAL FK columns —
        each further part is an FK column of the table reached before it.
        Examples (in `shipments`):

            ctx.related("order_id", "status")                # the order
            ctx.related("order_id.customer_id", "country")   # -> the customer
        """
        parts = [p.strip() for p in str(fk_path).split(".") if p.strip()]
        if not parts:
            raise RuntimeError('ctx.related: empty foreign key path')

        table_def = self._table
        keys = None
        parent_key = None
        for part in parts:
            col_def = next((c for c in table_def["columns"] if c["name"] == part), None)
            if col_def is None or not col_def.get("fk") or not col_def.get("fk_table"):
                raise RuntimeError(
                    f'ctx.related("{fk_path}", ...): "{part}" is not a foreign key column of {table_def["label"]}'
                )
            data = self._runner.data.get(table_def["label"], {})
            if part not in data:
                raise RuntimeError(f'ctx.related("{fk_path}", ...): column {table_def["label"]}.{part} is not generated yet')
            if keys is None:
                # First hop: our own FK values identify the records of the next table.
                keys = pd.Series(data[part])
            else:
                # Further hop: translate the keys collected so far into the FK
                # values of the table just reached.
                mapping = pd.Series(data[part].values, index=data[parent_key].values)
                mapping = mapping[~mapping.index.duplicated(keep="first")]
                keys = keys.map(mapping)
            parent_label, parent_key = col_def["fk_table"], col_def["fk_column"]
            next_table = self._runner.tables.get(parent_label)
            if next_table is None:
                raise RuntimeError(f'ctx.related("{fk_path}", ...): table {parent_label} is not part of this run')
            table_def = next_table

        final = self._runner.data.get(table_def["label"])
        if final is None or parent_key not in final or column not in final:
            raise RuntimeError(
                f'ctx.related("{fk_path}", "{column}"): values of {table_def["label"]}.{column} are not available (yet)'
            )
        mapping = pd.Series(final[column].values, index=final[parent_key].values)
        mapping = mapping[~mapping.index.duplicated(keep="first")]
        return pd.Series(keys).map(mapping)

    def lookup(self, name, column):
        """All (raw) values of a column of a lookup list (.lkp) — no drawing, no weighting."""
        values, _weights = self._runner.lookup_column(name, column)
        # A copy, so custom code can modify the list safely — the original list
        # lives in the runner's cache (see lookup_column).
        return list(values)

    def lookup_value(self, name, column):
        """
        ONE value per record from the consistently drawn list row — exactly the
        same mechanism as the built-in lookup generator (see
        Runner.lookup_indices): every column of the same list (including lookup
        generator columns and tables related via FK) reads the same row.
        Example: if the customer draws code "CH", then
        ctx.lookup_value("countries", "currency") on their order returns "CHF"
        from that same row.
        """
        n = self._runner.row_counts.get(self._table["label"])
        if n is None:
            raise RuntimeError('ctx.lookup_value(...): row count of this table is not known yet')
        values, _weights = self._runner.lookup_column(name, column)
        if not values:
            return pd.Series([""] * n, dtype=object)
        indices = self._runner.lookup_indices(self._table, name, n)
        arr = np.array(values, dtype=object)
        return pd.Series(arr[np.clip(indices, 0, len(arr) - 1)])


def indent_body(body):
    """Indents a method body so it can be compiled under a generated signature."""
    return "\n".join("    " + line for line in (body or "pass").split("\n"))


def os_path_stem(path):
    """File name of a path without its extension (for ctx.file_name)."""
    import os

    return os.path.splitext(os.path.basename(path))[0]


class FileContext:
    """
    The `ctx` object handed to a custom file generator's `write` method (see the
    .filegen editor). Besides facts about the run it offers the built-in writers
    as ready-made renderings, so a generator that only wraps the records — a
    header block in front of a CSV, say — does not have to re-implement them.
    """

    def __init__(self, runner, table, records, file_name):
        self._runner = runner
        self._table = table
        self.pd = pd
        self.np = np
        self.records = records
        self.file_name = file_name
        self.now = RUN_DT
        self.table_name = table.get("name", "")
        self.schema = (table.get("schema") or "").strip()
        self.label = table.get("label", "")

    @property
    def columns(self):
        """Names of the columns being written (hidden ones already dropped)."""
        return [c["name"] for c in self._table["columns"] if not c.get("hidden")]

    def log(self, *args):
        """Writes a message to the run log (the "Datenschmiede" output channel)."""
        emit("log", table=self._table["label"], message=" ".join(str(a) for a in args))

    def as_csv(self, df, **overrides):
        """
        The records as CSV text. Without arguments the table's own CSV settings
        apply; individual ones can be overridden (delimiter, include_header,
        decimal, quote_all).
        """
        cfg = {**(self._table["output"].get("csv") or {}), **overrides}
        return df.to_csv(
            index=False,
            sep=cfg.get("delimiter") or ";",
            decimal=cfg.get("decimal") or ".",
            header=cfg.get("include_header", True),
            quoting=csv_module.QUOTE_ALL if cfg.get("quote_all", True) else csv_module.QUOTE_MINIMAL,
            lineterminator="\n",
        )

    def as_json(self, df):
        """The records as a JSON document, using the table's JSON structure tab."""
        return self._runner.render_json(self._table, df)

    def as_xml(self, df):
        """The records as an XML document, using the table's XML structure tab."""
        return self._runner.render_xml(self._table, df)

    def as_fixed(self, df):
        """The records as fixed-length lines, using the table's record layout."""
        return self._runner.render_fixed(self._table, df)

    def as_excel(self, df, sheet_name=None):
        """The records as an Excel workbook — returns bytes, ready to be returned from write()."""
        import io as _io

        try:
            import openpyxl  # noqa: F401 — only needed as the pandas engine
        except ImportError:
            emit("error", code="missing-packages", packages=["openpyxl"],
                 message="Missing Python package: openpyxl (required for Excel output)")
            sys.exit(3)
        buffer = _io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sanitize_sheet_name(sheet_name or self.table_name), index=False)
        return buffer.getvalue()


def build_custom_functions(defn):
    """Builds generate/parse_params from the .tdgen code shipped in the plan."""
    namespace = {"np": np, "pd": pd}
    source = (
        f'def generate(params, n, ctx):\n{indent_body(defn.get("generate"))}\n\n'
        f'def parse_params(params):\n{indent_body(defn.get("parse_params") or "return params")}\n'
    )
    try:
        exec(compile(source, f'<tdgen:{defn["name"]}>', "exec"), namespace)
    except SyntaxError as err:
        raise RuntimeError(f'Custom generator "{defn["name"]}": syntax error in code ({err})') from err
    return namespace["generate"], namespace["parse_params"]


def sanitize_filename(name):
    """Replaces characters that are invalid in file names; never returns an empty name."""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip().strip(".")
    return name or "output"


def sanitize_path_segment(part):
    """Like sanitize_filename, but for individual folder segments (lets `.`/`..` through)."""
    if part in (".", ".."):
        return part
    return re.sub(r'[<>:"|?*\x00-\x1f]', "_", part).strip() or "_"


def sanitize_sheet_name(name):
    """
    Excel worksheet name: at most 31 characters, without the characters Excel
    forbids and never empty (otherwise openpyxl raises while writing).
    """
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", str(name or "")).strip().strip("'")
    return (cleaned or "Sheet1")[:31]


def parse_cell_ref(ref):
    """
    Splits an A1-style cell reference into its 0-based (row, column) offsets —
    the top-left corner the Excel table is written at. Anything unparseable
    falls back to A1.
    """
    match = re.fullmatch(r"\s*([A-Za-z]{1,3})\s*(\d{1,7})\s*", str(ref or ""))
    if not match:
        return 0, 0
    letters, digits = match.group(1).upper(), int(match.group(2))
    column = 0
    for char in letters:
        column = column * 26 + (ord(char) - ord("A") + 1)
    return max(0, digits - 1), max(0, column - 1)


def sanitize_xml_name(name):
    """
    Turns a configured name into a usable XML element/attribute name: invalid
    characters become underscores and a leading digit gets one prefixed.
    """
    cleaned = re.sub(r"[^\w.\-]", "_", str(name or ""), flags=re.UNICODE)
    if not cleaned:
        return "value"
    if not re.match(r"[A-Za-z_]", cleaned[0]):
        cleaned = "_" + cleaned
    return cleaned


def node_source_value(node, row):
    """The raw value a mapped leaf takes from one record (a column value, or its constant)."""
    if node.get("source_kind") == "constant":
        return node.get("source", "")
    return row.get(str(node.get("source") or ""), None)


def json_scalar(value, value_type):
    """
    Converts one mapped value to the JSON type configured for the node.
    `auto` keeps the generated value's own type (numbers stay numbers, dates
    are already formatted as text); a value that does not fit the requested
    type is written as text instead of failing the whole run.
    """
    if value is None or (isinstance(value, float) and value != value):
        # None and NaN (pandas' missing marker) both become JSON null.
        return None
    if hasattr(value, "item"):
        # numpy scalars (int64, bool_, float64) are not JSON-serializable.
        value = value.item()
    if value_type == "string":
        return str(value)
    if value_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("true", "1", "yes", "ja", "wahr")
    if value_type in ("number", "integer"):
        try:
            number = float(str(value).replace(",", ".")) if not isinstance(value, (int, float)) else float(value)
        except (TypeError, ValueError):
            return str(value)
        return int(number) if value_type == "integer" else number
    return value


def json_node_value(node, row):
    """Builds the JSON value of one structure node for one record (recursive)."""
    kind = node.get("kind")
    if kind == "object":
        return {
            sanitize_json_key(child, index): json_node_value(child, row)
            for index, child in enumerate(node.get("children") or [])
        }
    if kind == "array":
        # An array node writes one entry per child — the child names are
        # irrelevant there, only their values and order.
        return [json_node_value(child, row) for child in (node.get("children") or [])]
    return json_scalar(node_source_value(node, row), node.get("value_type") or "auto")


def sanitize_json_key(node, index):
    """Property name of a JSON node — never empty, so no record loses a field."""
    return str(node.get("name") or "").strip() or f"field_{index + 1}"


def xml_text(value):
    """Renders one mapped value as XML text (missing values become an empty element)."""
    if value is None or (isinstance(value, float) and value != value):
        return ""
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def fixed_text(value, decimal="."):
    """Renders one value as the text that goes into a fixed-length field."""
    if value is None or (isinstance(value, float) and value != value):
        return ""
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value)
    if decimal and decimal != "." and isinstance(value, float):
        text = text.replace(".", decimal)
    return text


def pad_fixed(text, field, truncate):
    """
    Pads (or cuts) one value to its field width. An over-long value is cut by
    default: leaving it would push every following field out of position, which
    is exactly what a fixed-length reader cannot cope with.
    """
    width = max(0, int(field.get("width") or 0))
    if width == 0:
        return ""
    pad = (str(field.get("pad") or " ") or " ")[0]
    if len(text) > width and truncate:
        # Right-aligned fields (typically numbers) keep their LAST characters —
        # cutting the low-order digits off a number would be far worse.
        text = text[-width:] if field.get("align") == "right" else text[:width]
    if field.get("align") == "right":
        return text.rjust(width, pad)
    return text.ljust(width, pad)


def xml_render_node(node, row, depth, step, newline, index=0):
    """Renders ONE structure node (object/array/value) of one record as XML text parts."""
    name = sanitize_xml_name(node.get("name") or f"field_{index + 1}")
    kind = node.get("kind")
    if kind == "object":
        return xml_element(name, node.get("children") or [], row, depth, step, newline)
    if kind == "array":
        # A repeating element: every child produces ONE element named after the
        # ARRAY node, not after the child. The child names are irrelevant here —
        # exactly as in JSON, where array entries are unnamed:
        #
        #   Tags [array]        ->  <Tags>gross</Tags>
        #     … [value]             <Tags>webshop</Tags>
        #     … [value]
        parts = []
        for child in node.get("children") or []:
            child_kind = child.get("kind")
            if child_kind == "object":
                # An object entry contributes its own children inside the
                # repeated element.
                parts.extend(xml_element(name, child.get("children") or [], row, depth, step, newline))
            elif child_kind == "array":
                # A nested list keeps its own name (that is what it is for).
                parts.extend(xml_render_node(child, row, depth, step, newline))
            else:
                text = xml_escape(xml_text(node_source_value(child, row)))
                parts.append(f"{step * depth}<{name}>{text}</{name}>{newline}")
        return parts
    text = xml_escape(xml_text(node_source_value(node, row)))
    return [f"{step * depth}<{name}>{text}</{name}>{newline}"]


def xml_element(name, children, row, depth, step, newline):
    """
    Renders one XML element with its children for one record: `attribute`
    children become attributes of this element, everything else a nested node.
    Returns the text parts in order (joined by the caller).
    """
    pad = step * depth
    attributes = ""
    inner = []
    for index, child in enumerate(children or []):
        if child.get("kind") == "attribute":
            attr_name = sanitize_xml_name(child.get("name") or f"attr_{index + 1}")
            attributes += f" {attr_name}={xml_quoteattr(xml_text(node_source_value(child, row)))}"
        else:
            inner.extend(xml_render_node(child, row, depth + 1, step, newline, index))

    if not inner:
        return [f"{pad}<{name}{attributes} />{newline}"]
    return [f"{pad}<{name}{attributes}>{newline}", *inner, f"{pad}</{name}>{newline}"]


class Runner:
    """Executes one plan: ordering, generation and writing of all tables."""

    def __init__(self, plan):
        self.plan = plan
        self.tables = {t["label"]: t for t in plan["tables"]}
        self.lookups = {l["name"]: l for l in plan.get("lookups", [])}
        self.custom = {}
        for defn in plan.get("custom_generators", []):
            self.custom[defn["name"]] = defn
        # Custom file generators (.filegen) — keyed by their logical name, as a
        # table's `format = "custom:<name>"` refers to them.
        self.file_generators = {}
        for defn in plan.get("file_generators", []):
            self.file_generators[defn["name"]] = defn
        # label -> {column name -> pd.Series} of the values generated so far
        self.data = {}
        # label -> row count
        self.row_counts = {}
        # (table_label, list_name) -> the lookup list row indices drawn per
        # record: this way every column drawing from the same list reads the
        # same row — across FK-related tables as well (see lookup_indices).
        self.lookup_row_indices = {}
        # list_name -> {"values": {column: [...]}, "weights": [...]} — the
        # column values/parsed weights extracted once per list, instead of
        # rebuilding them for every column that draws from the list.
        self._lookup_cache = {}

    # ------------------------------------------------------------------
    # Ordering
    # ------------------------------------------------------------------

    def table_dependencies(self, table):
        """Labels of the tables (in the plan) this table depends on."""
        deps = set()
        for column in table["columns"]:
            if column.get("fk") and column.get("fk_table") in self.tables and column["fk_table"] != table["label"]:
                deps.add(column["fk_table"])
            generator = column.get("generator")
            if generator:
                for ref in generator.get("table_refs", []):
                    if ref in self.tables and ref != table["label"]:
                        deps.add(ref)
        return deps

    def sorted_tables(self):
        """Topological sort of the tables; aborts with a clear message on cycles."""
        remaining = dict(self.tables)
        ordered = []
        while remaining:
            ready = [t for t in remaining.values() if not (self.table_dependencies(t) & set(remaining))]
            if not ready:
                fail("Circular dependency between tables: " + ", ".join(sorted(remaining)))
            for table in sorted(ready, key=lambda t: t["label"]):
                ordered.append(table)
                del remaining[table["label"]]
        return ordered

    def sorted_columns(self, table):
        """
        Column order within a table: the driving FK column is produced during
        row construction already; then column by column, combine columns only
        after the columns they reference, and custom generators last (their code
        can reach everything generated so far via ctx.column(...)).
        """
        columns = [c for c in table["columns"] if c["name"] != table.get("driving_fk_column")]

        def own_deps(column):
            generator = column.get("generator")
            if not generator:
                return set()
            deps = set(generator.get("own_column_refs", []))
            return deps

        remaining = {c["name"]: c for c in columns}
        # The declared column order as an index map, built once (calling
        # list.index per sort key would be quadratic).
        order_index = {c["name"]: i for i, c in enumerate(table["columns"])}
        ordered = []
        # First everything without own dependencies and without custom code,
        # then the rest.
        while remaining:
            ready = [
                c for c in remaining.values()
                if not (own_deps(c) & set(remaining))
            ]
            if not ready:
                fail(
                    f'Table {table["label"]}: circular column dependency between '
                    + ", ".join(sorted(remaining))
                )
            ready.sort(key=lambda c: (1 if (c.get("generator") or {}).get("id", "").startswith("custom:") else 0,
                                      order_index.get(c["name"], 0)))
            column = ready[0]
            ordered.append(column)
            del remaining[column["name"]]
        return ordered

    # ------------------------------------------------------------------
    # Lookup lists
    # ------------------------------------------------------------------

    def lookup_column(self, name, column):
        """Values and weights of one lookup list column (cached per list)."""
        lookup = self.lookups.get(name)
        if lookup is None:
            raise RuntimeError(f'Lookup list "{name}" was not found')
        if column not in lookup["columns"]:
            raise RuntimeError(f'Lookup list "{name}" has no column "{column}"')
        # Build column values and weights only once per list — every further
        # column drawing from the same list reads the cache.
        cache = self._lookup_cache.setdefault(name, {"values": {}, "weights": None})
        if column not in cache["values"]:
            index = lookup["columns"].index(column)
            cache["values"][column] = [
                row["values"][index] if index < len(row["values"]) else "" for row in lookup["rows"]
            ]
        if cache["weights"] is None:
            weights = []
            for row in lookup["rows"]:
                raw = str(row.get("weight", "")).strip().replace(",", ".")
                try:
                    weights.append(max(0.0, float(raw)))
                except ValueError:
                    weights.append(0.0)
            cache["weights"] = weights
        return cache["values"][column], cache["weights"]

    def lookup_indices(self, table, name, n):
        """
        The lookup list row indices drawn for a table — ONE row per record,
        read jointly by every column of the same list (e.g. code "DE" and
        currency "EUR" from the same row):

        1. If this table already has indices for the list (another column drew
           first), they are reused.
        2. Otherwise: if a table referenced via FK has indices for the list,
           they are taken over row-accurately (a join over the FK column) — so
           related records read the same list row across table boundaries.
        3. Otherwise a fresh weighted draw is made.
        """
        key = (table["label"], name)
        if key in self.lookup_row_indices:
            return self.lookup_row_indices[key]

        for column in table["columns"]:
            if not column.get("fk") or not column.get("fk_table"):
                continue
            parent_indices = self.lookup_row_indices.get((column["fk_table"], name))
            own_fk = self.data.get(table["label"], {}).get(column["name"])
            parent = self.data.get(column["fk_table"])
            if parent_indices is None or own_fk is None or parent is None:
                continue
            mapping = pd.Series(np.asarray(parent_indices), index=parent[column["fk_column"]].values)
            mapping = mapping[~mapping.index.duplicated(keep="first")]
            mapped = pd.Series(own_fk).map(mapping)
            if mapped.isna().any():
                # FK values without a match (should not happen) -> do not guess,
                # draw fresh below instead.
                break
            indices = mapped.to_numpy(dtype=np.int64)
            self.lookup_row_indices[key] = indices
            return indices

        lookup = self.lookups.get(name)
        if lookup is None:
            raise RuntimeError(f'Lookup list "{name}" was not found')
        count = len(lookup["rows"])
        if count == 0:
            indices = np.zeros(0, dtype=np.int64) if n == 0 else np.full(n, -1, dtype=np.int64)
            self.lookup_row_indices[key] = indices
            return indices
        _values, weights = self.lookup_column(name, lookup["columns"][0]) if lookup["columns"] else ([], [])
        total = sum(weights)
        p = [w / total for w in weights] if total > 0 else None
        indices = RNG.choice(count, size=n, p=p)
        self.lookup_row_indices[key] = indices
        return indices

    # ------------------------------------------------------------------
    # Generators
    # ------------------------------------------------------------------

    def generate_column(self, table, column, n):
        """Produces the `n` values of one column using its configured generator."""
        generator = column.get("generator")
        gen_id = (generator or {}).get("id", "")
        params = (generator or {}).get("params", {})

        if gen_id == "default":
            # The explicitly chosen per-data-type default — identical to a
            # column without any generator at all.
            return self.default_by_type(table, column, n)

        if gen_id == "foreign-key" or (column.get("fk") and not gen_id):
            ref_table, ref_column = column.get("fk_table"), column.get("fk_column")
            ref_values = self.data.get(ref_table, {}).get(ref_column)
            if ref_values is None:
                raise RuntimeError(
                    f'Column {table["label"]}.{column["name"]}: referenced values '
                    f'{ref_table}.{ref_column} are not available'
                )
            return pd.Series(RNG.choice(ref_values.to_numpy(), size=n))

        if gen_id == "random-int":
            low = int(params.get("min", "0"))
            high = int(params.get("max", "100"))
            return pd.Series(RNG.integers(low, high + 1, size=n))

        if gen_id == "random-float":
            low = float(str(params.get("min", "0")).replace(",", "."))
            high = float(str(params.get("max", "1")).replace(",", "."))
            decimals = int(params.get("decimals", "2") or "2")
            return pd.Series(np.round(RNG.uniform(low, high, size=n), decimals))

        if gen_id == "faker":
            provider = params.get("provider", "word")
            faker = get_faker(params.get("locale", "en_US"))
            method = getattr(faker, provider, None)
            if method is None:
                raise RuntimeError(
                    f'Column {table["label"]}.{column["name"]}: unknown Faker provider "{provider}"'
                )
            return pd.Series([method() for _ in range(n)], dtype=object)

        if gen_id == "lookup":
            # ONE list row is drawn per record (see lookup_indices) — every
            # column of the same list, including in FK-related tables, reads
            # that same row.
            list_name = params.get("list", "")
            values, _weights = self.lookup_column(list_name, params.get("column", ""))
            if not values:
                return pd.Series([""] * n, dtype=object)
            indices = self.lookup_indices(table, list_name, n)
            arr = np.array(values, dtype=object)
            return pd.Series(arr[np.clip(indices, 0, len(arr) - 1)])

        if gen_id == "combine":
            template = params.get("template", "")
            table_data = self.data[table["label"]]
            parts = re.split(r"(\{[^{}]+\})", template)
            pieces = []
            for part in parts:
                if not part:
                    continue
                if part.startswith("{") and part.endswith("}"):
                    ref = part[1:-1].strip()
                    if ref not in table_data:
                        raise RuntimeError(
                            f'Column {table["label"]}.{column["name"]}: combine template references '
                            f'"{ref}", which is not generated yet'
                        )
                    pieces.append(table_data[ref].astype(str).reset_index(drop=True))
                else:
                    pieces.append(pd.Series([part] * n, dtype=object))
            if not pieces:
                return pd.Series([""] * n, dtype=object)
            if len(pieces) == 1:
                return pieces[0]
            # Concatenate all parts in ONE pass, instead of repeatedly creating
            # intermediate series pairwise.
            return pieces[0].str.cat(pieces[1:])

        if gen_id.startswith("custom:"):
            name = gen_id[len("custom:"):]
            defn = self.custom.get(name)
            if defn is None:
                raise RuntimeError(
                    f'Column {table["label"]}.{column["name"]}: custom generator "{name}" was not found'
                )
            if "functions" not in defn:
                defn["functions"] = build_custom_functions(defn)
            generate, parse_params = defn["functions"]
            ctx = Context(self, table)
            parsed = parse_params(dict(params))
            series = generate(parsed, n, ctx)
            series = pd.Series(series).reset_index(drop=True)
            if len(series) != n:
                raise RuntimeError(
                    f'Custom generator "{name}" returned {len(series)} values, expected {n}'
                )
            return series

        # No (known) generator configured -> a sensible default per data type.
        return self.default_by_type(table, column, n)

    def default_by_type(self, table, column, n):
        """Default values per data type — also used for columns without a generator."""
        ctype = column.get("type", "string")
        name = column.get("name", "value")
        if ctype == "integer":
            # A running number — unique out of the box when used as a PK.
            return pd.Series(np.arange(1, n + 1, dtype=np.int64))
        if ctype in ("float", "decimal"):
            return pd.Series(np.round(RNG.uniform(0, 1000, size=n), 2))
        if ctype == "boolean":
            return pd.Series(RNG.integers(0, 2, size=n).astype(bool))
        if ctype == "uuid":
            # Vectorized UUID4 from random bytes (much faster than uuid.uuid4
            # per row). The complete buffer is hex-encoded in ONE C call
            # (instead of a Python callback per row, as apply_along_axis would
            # do) and then split into 32-character chunks.
            b = RNG.integers(0, 256, size=(n, 16), dtype=np.uint8)
            b[:, 6] = (b[:, 6] & 0x0F) | 0x40
            b[:, 8] = (b[:, 8] & 0x3F) | 0x80
            full = b.tobytes().hex()
            return pd.Series(
                [
                    f"{full[i:i + 8]}-{full[i + 8:i + 12]}-{full[i + 12:i + 16]}-{full[i + 16:i + 20]}-{full[i + 20:i + 32]}"
                    for i in range(0, 32 * n, 32)
                ],
                dtype=object,
            )
        if ctype == "date":
            start = RUN_DT - timedelta(days=365)
            offsets = RNG.integers(0, 365, size=n)
            base = pd.Timestamp(start.date())
            return pd.Series(base + pd.to_timedelta(offsets, unit="D")).dt.normalize()
        if ctype == "datetime":
            seconds = RNG.integers(0, 365 * 24 * 3600, size=n)
            base = pd.Timestamp(RUN_DT) - pd.Timedelta(days=365)
            return pd.Series(base + pd.to_timedelta(seconds, unit="s"))
        if ctype == "time":
            seconds = RNG.integers(0, 24 * 3600, size=n)
            return pd.Series(pd.to_timedelta(seconds, unit="s"))
        if ctype == "email":
            nums = np.arange(1, n + 1)
            return pd.Series([f"user{i}@example.com" for i in nums], dtype=object)
        if ctype == "json":
            return pd.Series(["{}"] * n, dtype=object)
        # string / text / anything unknown
        return pd.Series([f"{name}_{i}" for i in range(1, n + 1)], dtype=object)

    # ------------------------------------------------------------------
    # Table run
    # ------------------------------------------------------------------

    def run_table(self, table):
        """Generates all rows of one table and returns their count."""
        label = table["label"]
        self.data[label] = {}

        driving_lookup = table.get("driving_lookup")
        if driving_lookup:
            # A leading lookup list: exactly one record per list row, each row
            # used once and in list order. Seeding the drawn row indices with
            # arange means every lookup generator on this table (and on tables
            # related to it by FK) reads that same row — so a predefined list of
            # IDs becomes one record each, with all its other columns matching.
            lookup = self.lookups.get(driving_lookup)
            if lookup is None:
                raise RuntimeError(
                    f'Table {label}: leading lookup list "{driving_lookup}" was not found'
                )
            n = len(lookup["rows"])
            self.lookup_row_indices[(label, driving_lookup)] = np.arange(n, dtype=np.int64)
        elif table.get("driving_fk"):
            driving = table["driving_fk"]
            parent_label, parent_column = driving["table"], driving["column"]
            parent_values = self.data.get(parent_label, {}).get(parent_column)
            if parent_values is None:
                raise RuntimeError(
                    f'Table {label}: referenced values {parent_label}.{parent_column} are not available'
                )
            lo, hi = int(table["records"]["min"]), int(table["records"]["max"])
            counts = RNG.integers(lo, hi + 1, size=len(parent_values))
            n = int(counts.sum())
            # The driving FK column is produced right during row construction:
            # every record of the referenced table gets `counts` related
            # records.
            self.data[label][table["driving_fk_column"]] = pd.Series(
                np.repeat(parent_values.to_numpy(), counts)
            )
        else:
            n = int(table["records"])

        self.row_counts[label] = n

        for column in self.sorted_columns(table):
            self.data[label][column["name"]] = self.generate_column(table, column, n).reset_index(drop=True)
            # Column-by-column progress for the run log (output channel).
            emit("column_done", table=label, column=column["name"], records=n)

        return n

    # ------------------------------------------------------------------
    # Output
    # ------------------------------------------------------------------

    def resolve_output_dir(self):
        """
        Resolves the plan's output folder: the template's `{…}` variables
        (date/time/timestamp/project name) are substituted and every path
        segment is sanitized; relative paths are resolved against the folder of
        the project file. Empty -> `output`.
        """
        import os

        template = (self.plan.get("output_path") or "").strip() or "output"

        def replace(match):
            token = match.group(1).strip()
            if token == "date":
                return RUN_DT.strftime("%Y%m%d")
            if token == "time":
                return RUN_DT.strftime("%H%M%S")
            if token == "datetime":
                return RUN_DT.strftime("%Y%m%d_%H%M%S")
            if token == "timestamp":
                return str(int(RUN_DT.timestamp()))
            if token == "project":
                return self.plan.get("project_name", "")
            return ""

        resolved = re.sub(r"\{([^{}]+)\}", replace, template)
        # Preserve separators (and a leading drive such as `C:`), sanitizing
        # each segment individually.
        parts = re.split(r"([\\/]+)", resolved)
        cleaned = "".join(
            part if re.fullmatch(r"[\\/]+", part) or re.fullmatch(r"[A-Za-z]:", part) else sanitize_path_segment(part)
            for part in parts
            if part
        )
        return os.path.normpath(os.path.join(self.plan.get("project_dir", "."), cleaned or "output"))

    def resolve_template(self, table, n, df, template):
        """
        Substitutes the `{…}` variables of a file name / sheet name template
        (see FILE_NAME_VARIABLES in src/table/model.ts); unknown tokens vanish.
        """

        def replace(match):
            token = match.group(1).strip()
            if token == "date":
                return RUN_DT.strftime("%Y%m%d")
            if token == "time":
                return RUN_DT.strftime("%H%M%S")
            if token == "datetime":
                return RUN_DT.strftime("%Y%m%d_%H%M%S")
            if token == "timestamp":
                return str(int(RUN_DT.timestamp()))
            if token == "schema":
                return (table.get("schema") or "").strip()
            if token == "table":
                return table["name"]
            if token == "records":
                return str(n)
            if token.startswith("column:"):
                cname = token[len("column:"):].strip()
                if cname in df.columns and len(df) > 0:
                    return str(df[cname].iloc[0])
                return ""
            return ""

        return re.sub(r"\{([^{}]+)\}", replace, template)

    def resolve_filename(self, table, n, df):
        """Resolves a table's file name template (without extension) for this run."""
        template = (table["output"].get("file_name") or "").strip()
        if not template:
            schema = (table.get("schema") or "").strip()
            template = f"{schema}_{table['name']}" if schema else table["name"]
        return sanitize_filename(self.resolve_template(table, n, df, template))

    def format_dataframe(self, table, df, cfg=None):
        """
        Renders date/time columns as text per the configured formats. `cfg` is
        the file type's own settings block — every format brings its own
        date/timestamp format; without one (the preview grid) the CSV settings
        apply.
        """
        if cfg is None:
            cfg = table["output"].get("csv", {})
        date_fmt = cfg.get("date_format") or "%Y-%m-%d"
        datetime_fmt = cfg.get("datetime_format") or "%Y-%m-%d %H:%M:%S"
        # A shallow copy suffices: below, only whole columns are REASSIGNED
        # (which leaves the original untouched) — a deep copy would needlessly
        # double the memory footprint on large runs.
        out = df.copy(deep=False)
        for column in table["columns"]:
            cname, ctype = column["name"], column.get("type")
            if cname not in out.columns:
                continue
            try:
                if ctype == "date":
                    out[cname] = pd.to_datetime(out[cname]).dt.strftime(date_fmt)
                elif ctype == "datetime":
                    out[cname] = pd.to_datetime(out[cname]).dt.strftime(datetime_fmt)
                elif ctype == "time":
                    series = out[cname]
                    if pd.api.types.is_timedelta64_dtype(series):
                        base = pd.Timestamp("1970-01-01") + series
                        out[cname] = base.dt.strftime("%H:%M:%S")
            except (ValueError, TypeError):
                # Leave values that cannot be read as a date/time (e.g. coming
                # from a custom generator) unchanged.
                pass
        return out

    def visible_frame(self, table, out):
        """
        Drops the hidden columns (hidden = true) — they are generated and
        available up to this point (as an FK source for other tables, for
        {column:...} file names) and are merely not written to the file.
        """
        visible = [c["name"] for c in table["columns"] if not c.get("hidden") and c["name"] in out.columns]
        return out[visible]

    def write_output(self, table, df, n):
        """
        Writes one table in its configured file type and returns the path of the
        file created — or None for the "temp" file type, which generates the
        records without writing anything.
        """
        import os

        fmt = (table["output"].get("format") or "csv").strip().lower()
        if fmt == "temp":
            # Temporary table: the records stay in memory for other tables to
            # reference (FK, ctx.table(...)); nothing lands on disk.
            return None
        os.makedirs(self.out_dir, exist_ok=True)
        if fmt.startswith("custom:"):
            return self.write_custom(table, df, n, fmt[len("custom:"):])
        if fmt == "xlsx":
            return self.write_xlsx(table, df, n)
        if fmt == "json":
            return self.write_json(table, df, n)
        if fmt == "xml":
            return self.write_xml(table, df, n)
        if fmt == "fixed":
            return self.write_fixed(table, df, n)
        return self.write_csv(table, df, n)

    def output_path(self, table, n, out, extension):
        """Full path of the file to write, from the resolved file name template."""
        import os

        return os.path.join(self.out_dir, self.resolve_filename(table, n, out) + "." + extension)

    def write_csv(self, table, df, n):
        """Writes one table as CSV and returns the path of the file created."""
        csv_cfg = table["output"].get("csv", {})
        out = self.format_dataframe(table, df, csv_cfg)
        path = self.output_path(table, n, out, "csv")
        out = self.visible_frame(table, out)
        out.to_csv(
            path,
            index=False,
            sep=csv_cfg.get("delimiter") or ";",
            decimal=csv_cfg.get("decimal") or ".",
            header=csv_cfg.get("include_header", True),
            quoting=csv_module.QUOTE_ALL if csv_cfg.get("quote_all", True) else csv_module.QUOTE_MINIMAL,
            encoding=csv_cfg.get("encoding") or "utf-8",
        )
        return path

    def write_xlsx(self, table, df, n):
        """
        Writes one table as an Excel workbook (via pandas/openpyxl) and returns
        the path of the file created: the sheet name comes from its own {…}
        template, the table is placed at the configured start cell, and the
        header row can optionally be frozen and given an auto filter.
        """
        try:
            import openpyxl  # noqa: F401 — only needed as the pandas engine
            from openpyxl.utils import get_column_letter
        except ImportError:
            emit("error", code="missing-packages", packages=["openpyxl"],
                 message="Missing Python package: openpyxl (required for Excel output)")
            sys.exit(3)

        cfg = table["output"].get("xlsx", {})
        out = self.format_dataframe(table, df, cfg)
        path = self.output_path(table, n, out, "xlsx")
        out = self.visible_frame(table, out)

        header = bool(cfg.get("include_header", True))
        start_row, start_col = parse_cell_ref(cfg.get("start_cell"))
        sheet_name = sanitize_sheet_name(self.resolve_template(table, n, out, cfg.get("sheet_name") or "{table}"))

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            out.to_excel(writer, sheet_name=sheet_name, index=False, header=header,
                         startrow=start_row, startcol=start_col)
            sheet = writer.sheets[sheet_name]
            first_data_row = start_row + (2 if header else 1)
            last_row = start_row + (1 if header else 0) + len(out)
            last_col = start_col + max(1, len(out.columns))
            if header and len(out.columns) > 0:
                if cfg.get("freeze_header", True):
                    # Everything above the first data row stays visible while
                    # scrolling (the columns left of the table are not frozen —
                    # a start cell merely offsets the table, it does not make
                    # those columns a row header).
                    sheet.freeze_panes = f"A{first_data_row}"
                if cfg.get("auto_filter"):
                    sheet.auto_filter.ref = (
                        f"{get_column_letter(start_col + 1)}{start_row + 1}"
                        f":{get_column_letter(last_col)}{max(last_row, start_row + 1)}"
                    )
            if cfg.get("auto_fit_columns", True):
                for index, name in enumerate(out.columns):
                    # openpyxl has no real auto-fit (that needs Excel's own text
                    # metrics) — the widest rendered value is a good enough
                    # approximation, capped so a long text column does not push
                    # everything else off screen.
                    widest = max([len(str(name)) if header else 0]
                                 + [len(str(v)) for v in out[name].head(1000)] + [1])
                    sheet.column_dimensions[get_column_letter(start_col + index + 1)].width = min(60, widest + 2)
        return path

    def write_json(self, table, df, n):
        """Writes one table as JSON (or JSON Lines) per its target structure."""
        cfg = table["output"].get("json", {})
        out = self.format_dataframe(table, df, cfg)
        path = self.output_path(table, n, out, "json")
        text = self.render_json(table, out)
        with open(path, "w", encoding=cfg.get("encoding") or "utf-8", newline="") as handle:
            handle.write(text)
        return path

    def write_xml(self, table, df, n):
        """Writes one table as XML per its target structure."""
        cfg = table["output"].get("xml", {})
        out = self.format_dataframe(table, df, cfg)
        path = self.output_path(table, n, out, "xml")
        text = self.render_xml(table, out)
        with open(path, "w", encoding=cfg.get("encoding") or "utf-8", newline="") as handle:
            handle.write(text)
        return path

    def write_fixed(self, table, df, n):
        """Writes one table as a fixed-length (flat) file per its field layout."""
        cfg = table["output"].get("fixed", {})
        out = self.format_dataframe(table, df, cfg)
        path = self.output_path(table, n, out, "txt")
        text = self.render_fixed(table, out)
        # newline="" keeps the line ending exactly as rendered — otherwise
        # Python would translate "\n" to the platform default and a file meant
        # to be LF would silently become CRLF on Windows.
        with open(path, "w", encoding=cfg.get("encoding") or "utf-8", newline="") as handle:
            handle.write(text)
        return path

    # ------------------------------------------------------------------
    # Custom file generators (.filegen)
    # ------------------------------------------------------------------

    def custom_writer(self, name):
        """Compiles a `.filegen` write method from the code shipped in the plan (once per run)."""
        defn = self.file_generators.get(name)
        if defn is None:
            raise RuntimeError(f'File generator "{name}" was not found')
        if "function" not in defn:
            namespace = {"np": np, "pd": pd}
            source = f'def write(df, ctx):\n{indent_body(defn.get("write"))}\n'
            try:
                exec(compile(source, f"<filegen:{name}>", "exec"), namespace)
            except SyntaxError as err:
                raise RuntimeError(f'File generator "{name}": syntax error in code ({err})') from err
            defn["function"] = namespace["write"]
        return defn["function"]

    def write_custom(self, table, df, n, name):
        """
        Hands the finished records to a custom file generator and writes what it
        returns: `str` with the table's encoding, `bytes` unchanged (that is how
        a binary format such as Excel gets out).
        """
        defn = self.file_generators.get(name)
        if defn is None:
            raise RuntimeError(
                f'Table {table["label"]}: file generator "{name}" was not found'
            )
        write = self.custom_writer(name)
        # The frame the generator sees is the one the built-in writers would see:
        # dates formatted, hidden columns dropped.
        out = self.visible_frame(table, self.format_dataframe(table, df, table["output"].get("csv", {})))
        extension = (defn.get("extension") or "txt").strip().lstrip(".") or "txt"
        path = self.output_path(table, n, out, extension)

        ctx = FileContext(self, table, n, os_path_stem(path))
        result = write(out, ctx)
        if isinstance(result, (bytes, bytearray)):
            with open(path, "wb") as handle:
                handle.write(result)
        else:
            encoding = table["output"].get("csv", {}).get("encoding") or "utf-8"
            with open(path, "w", encoding=encoding, newline="") as handle:
                handle.write("" if result is None else str(result))
        return path

    # ------------------------------------------------------------------
    # Fixed length: field layout
    # ------------------------------------------------------------------

    def fixed_fields(self, table, cfg):
        """
        The configured field layout — or, while none has been set up, one field
        per written column with the default width (the counterpart of
        fixedFieldsFromColumns in src/table/model.ts).
        """
        fields = cfg.get("fields") or []
        if fields:
            return fields
        return [
            {"column": c["name"], "width": 20, "align": "left", "pad": " "}
            for c in table["columns"]
            if not c.get("hidden") and str(c.get("name") or "").strip()
        ]

    def render_fixed(self, table, out):
        """Renders the (already formatted) records as fixed-length lines."""
        cfg = table["output"].get("fixed", {})
        fields = self.fixed_fields(table, cfg)
        truncate = bool(cfg.get("truncate", True))
        newline = "\r\n" if str(cfg.get("line_ending") or "lf").lower() == "crlf" else "\n"
        decimal = cfg.get("decimal") or "."

        lines = []
        if cfg.get("include_header"):
            # The header carries the column names through exactly the same
            # padding, so it lines up with the data below it.
            lines.append("".join(pad_fixed(str(f.get("column") or ""), f, truncate) for f in fields))
        for row in out.to_dict("records"):
            lines.append(
                "".join(pad_fixed(fixed_text(row.get(str(f.get("column") or "")), decimal), f, truncate) for f in fields)
            )
        return "".join(line + newline for line in lines)

    # ------------------------------------------------------------------
    # JSON/XML: target structure + mapping
    # ------------------------------------------------------------------

    def structure_nodes(self, table, cfg, leaf_kind="value"):
        """
        The configured target structure of a table — or, while none has been set
        up, a flat fallback with one mapped leaf per written column, so that
        switching a table to JSON/XML produces sensible output right away (the
        counterpart of structureFromColumns in src/table/model.ts).
        """
        nodes = cfg.get("nodes") or []
        if nodes:
            return nodes
        return [
            {"name": c["name"], "kind": leaf_kind, "value_type": "auto",
             "source_kind": "column", "source": c["name"], "children": []}
            for c in table["columns"]
            if not c.get("hidden") and str(c.get("name") or "").strip()
        ]

    def render_json(self, table, out):
        """Renders the (already formatted) records as a JSON document."""
        cfg = table["output"].get("json", {})
        nodes = self.structure_nodes(table, cfg)
        # JSON and XML are record-shaped formats: unlike CSV they cannot be
        # written straight from the vectorized frame, so the values are taken
        # row by row here.
        records = [json_node_value({"kind": "object", "children": nodes}, row) for row in out.to_dict("records")]

        indent = max(0, int(cfg.get("indent") or 0))
        ensure_ascii = bool(cfg.get("ascii_only"))
        if cfg.get("json_lines"):
            # JSON Lines: one record per line, always compact — an indented
            # object would span several lines and break the format.
            return "".join(json.dumps(r, ensure_ascii=ensure_ascii) + "\n" for r in records)
        root_name = str(cfg.get("root_name") or "").strip()
        document = {root_name: records} if root_name else records
        return json.dumps(document, ensure_ascii=ensure_ascii, indent=indent or None) + "\n"

    def render_xml(self, table, out):
        """Renders the (already formatted) records as an XML document."""
        cfg = table["output"].get("xml", {})
        nodes = self.structure_nodes(table, cfg, leaf_kind="value")
        root = sanitize_xml_name(cfg.get("root_element") or "rows")
        record = sanitize_xml_name(cfg.get("record_element") or "row")
        indent = max(0, int(cfg.get("indent") or 0))
        step = " " * indent
        newline = "\n" if indent else ""

        parts = []
        if cfg.get("declaration", True):
            parts.append(f'<?xml version="1.0" encoding="{cfg.get("encoding") or "utf-8"}"?>' + (newline or "\n"))
        parts.append(f"<{root}>{newline}")
        for row in out.to_dict("records"):
            parts.extend(xml_element(record, nodes, row, 1, step, newline))
        parts.append(f"</{root}>\n")
        return "".join(parts)

    def render_preview_text(self, table, df):
        """
        The preview records as the configured file type would write them — for
        JSON/XML and fixed length, where the document itself (or the column
        alignment) is what needs checking; `None` for CSV/Excel, whose preview
        is the value grid.
        """
        fmt = (table["output"].get("format") or "csv").strip().lower()
        if fmt == "temp":
            # Nothing is written for a temporary table, so there is no file to show.
            return None
        if fmt == "json":
            return self.render_json(table, self.format_dataframe(table, df, table["output"].get("json", {})))
        if fmt == "xml":
            return self.render_xml(table, self.format_dataframe(table, df, table["output"].get("xml", {})))
        if fmt == "fixed":
            return self.render_fixed(table, self.format_dataframe(table, df, table["output"].get("fixed", {})))
        if fmt == "csv":
            csv_cfg = table["output"].get("csv", {})
            out = self.visible_frame(table, self.format_dataframe(table, df, csv_cfg))
            return out.to_csv(
                index=False,
                sep=csv_cfg.get("delimiter") or ";",
                decimal=csv_cfg.get("decimal") or ".",
                header=csv_cfg.get("include_header", True),
                quoting=csv_module.QUOTE_ALL if csv_cfg.get("quote_all", True) else csv_module.QUOTE_MINIMAL,
                lineterminator="\n",
            )
        if fmt.startswith("custom:"):
            # Run the custom writer over the preview records so its output can be
            # checked without writing anything.
            name = fmt[len("custom:"):]
            if name not in self.file_generators:
                return None
            out = self.visible_frame(table, self.format_dataframe(table, df, table["output"].get("csv", {})))
            result = self.custom_writer(name)(out, FileContext(self, table, len(out), "preview"))
            if isinstance(result, (bytes, bytearray)):
                # A binary format has nothing readable to show — report its size
                # instead of dumping raw bytes into the dialog.
                return f"<{len(result)} bytes>"
            return "" if result is None else str(result)
        return None

    def run(self):
        """Runs the whole plan: every table in dependency order, then the output."""
        ordered = self.sorted_tables()
        preview = self.plan.get("preview")
        self.out_dir = self.resolve_output_dir()
        emit("start", tables=len(ordered))
        files = []
        for index, table in enumerate(ordered):
            emit("table_start", table=table["label"], index=index, total=len(ordered))
            n = self.run_table(table)
            # Keep the column order of the table definition.
            df = pd.DataFrame({c["name"]: self.data[table["label"]][c["name"]] for c in table["columns"]})
            if preview:
                # Preview mode: write nothing; only the target table is
                # reported back at the end.
                if table["label"] == preview.get("table"):
                    limited = df.head(int(preview.get("limit", 20)))
                    out = self.format_dataframe(table, limited)
                    emit(
                        "preview",
                        table=table["label"],
                        columns=[c["name"] for c in table["columns"]],
                        rows=[["" if v is None else str(v) for v in row] for row in out.itertuples(index=False)],
                        # For JSON/XML the mapping tab shows the rendered
                        # document itself instead of a value grid.
                        text=self.render_preview_text(table, limited),
                    )
                continue
            path = self.write_output(table, df, n)
            if path is not None:
                files.append({"table": table["label"], "file": path, "records": n})
            # table_done is reported either way, so the progress indicator and
            # the record counts of the ER diagram stay complete.
            emit("table_done", table=table["label"], file=path or "", records=n, index=index, total=len(ordered))
        if not preview:
            emit("done", files=files, output_dir=self.out_dir)


def main():
    """Entry point: reads the plan file passed as an argument and runs it."""
    if len(sys.argv) < 2:
        fail("Usage: generate.py <plan.json>")
    try:
        with open(sys.argv[1], encoding="utf-8") as handle:
            plan = json.load(handle)
    except (OSError, json.JSONDecodeError) as err:
        fail(f"Unable to read plan file: {err}")
    try:
        Runner(plan).run()
    except RuntimeError as err:
        # Include the traceback even for "expected" errors — for failures in
        # custom generator code it points at the offending line
        # (<tdgen:name> frames), see the output channel in the extension host.
        fail(err, traceback=traceback.format_exc())
    except Exception as err:  # noqa: BLE001 — report every unexpected exception cleanly as an event
        fail(f"{type(err).__name__}: {err}", traceback=traceback.format_exc())


if __name__ == "__main__":
    main()
